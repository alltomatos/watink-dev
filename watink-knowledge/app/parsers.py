"""Extração de texto plano de arquivos de fontes da Base de Conhecimento.

`extract_text(filename, data)` escolhe o parser pelo sufixo do nome. Nunca levanta
por conteúdo ruim: retorna "" e deixa o worker marcar a fonte como sem texto.
Suporta pdf/docx/xlsx/csv e qualquer texto (txt/md/desconhecido via decode utf-8).
"""

import csv
import io
import logging
import os

log = logging.getLogger("parsers")


def _extract_pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    parts = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text:
            parts.append(text)
    return "\n".join(parts)


def _extract_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    parts = [p.text for p in doc.paragraphs if p.text]
    return "\n".join(parts)


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    parts.append("\t".join(cells))
    finally:
        wb.close()
    return "\n".join(parts)


def _extract_csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="ignore")
    parts = []
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        cells = [c for c in row if c and c.strip()]
        if cells:
            parts.append("\t".join(cells))
    return "\n".join(parts)


def extract_text(filename: str, data: bytes) -> str:
    """Extrai texto plano de `data` escolhendo o parser pelo sufixo de `filename`.

    Retorna string vazia se nada for extraído ou se o parser falhar — nunca levanta.
    """
    ext = os.path.splitext(filename or "")[1].lower()
    try:
        if ext == ".pdf":
            return _extract_pdf(data)
        if ext == ".docx":
            return _extract_docx(data)
        if ext == ".xlsx":
            return _extract_xlsx(data)
        if ext == ".csv":
            return _extract_csv(data)
        # .txt / .md / outros: decode plano
        return data.decode("utf-8", errors="ignore")
    except Exception:
        log.warning("falha ao extrair texto de %r (ext=%s)", filename, ext, exc_info=True)
        return ""
